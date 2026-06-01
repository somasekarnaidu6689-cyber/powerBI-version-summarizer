import json
from pathlib import Path
from datetime import datetime

# ── PDF ───────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, HRFlowable, Image
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER

# ── Excel ─────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
#  PDF EXPORT
# ═══════════════════════════════════════════════════════════════

def export_pdf(
    model_results: dict,
    report_results: dict,
    layout_results: dict,
    ai_summary: dict,
    old_path: str,
    new_path: str,
    output_path: Path
) -> Path:
    """Generates a structured PDF report from diff results."""

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title="Power BI Diff Report",
        author="pbix_diff Tool"
    )

    styles = _build_styles()
    story  = []

    # ── Cover ─────────────────────────────────────────────────
    logo_path = Path(__file__).parent.parent / 'output' / 'assets' / 'logo.png'
    if logo_path.exists():
        logo = Image(str(logo_path), width=4*cm, height=4*cm)
        logo.hAlign = 'LEFT'
        story.append(logo)
        story.append(Spacer(1, 0.3*cm))

    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("⚡ Power BI Version Comparison", styles["title"]))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["meta"]))
    story.append(Paragraph(f"Old: {old_path}", styles["meta"]))
    story.append(Paragraph(f"New: {new_path}", styles["meta"]))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1D4F55"), spaceAfter=12))

    # ── Summary Table ─────────────────────────────────────────
    story.append(Paragraph("Summary", styles["h1"]))
    summary_data = [
        ["Category", "Metric", "Count"],
        ["SemanticModel", "Tables Added",      str(len(model_results.get("tables_added", [])))],
        ["SemanticModel", "Tables Removed",    str(len(model_results.get("tables_removed", [])))],
        ["SemanticModel", "Columns Added",     str(len(model_results.get("columns_added", [])))],
        ["SemanticModel", "Columns Removed",   str(len(model_results.get("columns_removed", [])))],
        ["SemanticModel", "Columns Modified",  str(len(model_results.get("columns_modified", [])))],
        ["SemanticModel", "Measures Changed",  str(len(model_results.get("measures_modified", [])))],
        ["SemanticModel", "Partitions Changed",str(len(model_results.get("partitions_changed", [])))],
        ["Report",        "Pages Added",       str(len(report_results.get("pages_added", [])))],
        ["Report",        "Pages Removed",     str(len(report_results.get("pages_removed", [])))],
        ["Report",        "Pages Changed",     str(len(report_results.get("pages_changed", [])))],
    ]
    story.append(_make_table(summary_data, col_widths=[5*cm, 7*cm, 3*cm]))
    story.append(Spacer(1, 0.5*cm))

    # ── SemanticModel Section ─────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("SemanticModel Changes", styles["h1"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#2a3348"), spaceAfter=8))

    # Tables Added
    if model_results.get("tables_added"):
        story.append(Paragraph("Tables Added", styles["h2"]))
        for t in model_results["tables_added"]:
            is_auto = "LocalDateTable" in t["name"] or "DateTableTemplate" in t["name"]
            label   = f"[auto-generated date table] {t['name']}" if is_auto else t["name"]
            story.append(Paragraph(f"+ {label}", styles["added"]))
            if t.get("columns"):
                story.append(Paragraph(f"  Columns: {', '.join(t['columns'])}", styles["detail"]))
            if t.get("measures"):
                for m in t["measures"]:
                    story.append(Paragraph(f"  Measure: {m['name']}", styles["detail"]))
                    if m.get("dax"):
                        story.append(Paragraph(f"  DAX: {m['dax'][:200]}", styles["code"]))
            story.append(Spacer(1, 0.2*cm))

    # Tables Removed
    if model_results.get("tables_removed"):
        story.append(Paragraph("Tables Removed", styles["h2"]))
        for t in model_results["tables_removed"]:
            story.append(Paragraph(f"- {t['name']}", styles["removed"]))

    # Columns
    if model_results.get("columns_added") or model_results.get("columns_removed") or model_results.get("columns_modified"):
        story.append(Paragraph("Column Changes", styles["h2"]))
        col_data = [["Table", "Column", "Status", "Detail"]]
        for c in model_results.get("columns_added", []):
            col_data.append([c["table"], c["column"], "+ Added", c.get("dataType", "")])
        for c in model_results.get("columns_removed", []):
            col_data.append([c["table"], c["column"], "- Removed", ""])
        for c in model_results.get("columns_modified", []):
            detail = "; ".join(f"{k}: {v['old']} → {v['new']}" for k, v in c["changes"].items())
            col_data.append([c["table"], c["column"], "~ Modified", detail])
        story.append(_make_table(col_data, col_widths=[4*cm, 4*cm, 2.5*cm, 6.5*cm]))
        story.append(Spacer(1, 0.3*cm))

    # Measures
    if model_results.get("measures_modified"):
        story.append(Paragraph("Measure / DAX Changes", styles["h2"]))
        for m in model_results["measures_modified"]:
            story.append(Paragraph(f"{m['measure']} [{m['status'].upper()}]", styles["measure_name"]))
            if m["status"] == "modified":
                story.append(Paragraph("Old DAX:", styles["dax_label"]))
                story.append(Paragraph(str(m.get("old_dax") or ""), styles["code"]))
                story.append(Paragraph("New DAX:", styles["dax_label"]))
                story.append(Paragraph(str(m.get("new_dax") or ""), styles["code"]))
            elif m["status"] == "added":
                story.append(Paragraph(str(m.get("new_dax") or ""), styles["code"]))
            else:
                story.append(Paragraph(str(m.get("old_dax") or ""), styles["code"]))
            story.append(Spacer(1, 0.2*cm))

    # Partitions
    if model_results.get("partitions_changed"):
        story.append(Paragraph("Partition / Query Changes", styles["h2"]))
        for p in model_results["partitions_changed"]:
            story.append(Paragraph(f"{p['table']} / {p['partition']} [{p['status']}]", styles["measure_name"]))
            if p["status"] == "modified":
                story.append(Paragraph(f"Old ({p.get('old_mode')}): {str(p.get('old_source') or '')[:300]}", styles["code"]))
                story.append(Paragraph(f"New ({p.get('new_mode')}): {str(p.get('new_source') or '')[:300]}", styles["code"]))
            story.append(Spacer(1, 0.2*cm))

    # ── Report Section ────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("Report Changes", styles["h1"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#2a3348"), spaceAfter=8))

    # Pages Changed
    for pg in report_results.get("pages_changed", []):
        story.append(Paragraph(f"Page: {pg['display_name']} [CHANGED]", styles["h2"]))
        for v in pg.get("visuals_changed", []):
            story.append(Paragraph(
                f"  Visual [{v['status']}]: {v['visual_type']} ({v['visual_id']})",
                styles["added"] if v["status"] == "added" else
                styles["removed"] if v["status"] == "removed" else styles["modified"]
            ))
            if v.get("fields"):
                story.append(Paragraph(f"    Fields: {', '.join(v['fields'])}", styles["detail"]))
            ch = v.get("changes", {})
            if ch.get("visual_type"):
                story.append(Paragraph(
                    f"    Chart: {ch['visual_type']['old']} → {ch['visual_type']['new']}",
                    styles["detail"]
                ))
            if ch.get("field_swap"):
                rem = ", ".join(ch["field_swap"].get("removed", []))
                add = ", ".join(ch["field_swap"].get("added", []))
                story.append(Paragraph(f"    Fields removed: {rem}", styles["detail"]))
                story.append(Paragraph(f"    Fields added:   {add}", styles["detail"]))
        story.append(Spacer(1, 0.3*cm))

    # Pages Added
    for pg in report_results.get("pages_added", []):
        story.append(Paragraph(f"Page Added: {pg['display_name']}", styles["h2"]))
        for v in pg.get("visuals", []):
            story.append(Paragraph(
                f"  + {v['visual_type']} ({v['visual_id']})", styles["added"]
            ))
            if v.get("fields"):
                story.append(Paragraph(f"    Fields: {', '.join(v['fields'])}", styles["detail"]))
        story.append(Spacer(1, 0.3*cm))

    # Pages Removed
    for pg in report_results.get("pages_removed", []):
        story.append(Paragraph(f"Page Removed: {pg['display_name']}", styles["h2"]))
        for v in pg.get("visuals", []):
            story.append(Paragraph(
                f"  - {v['visual_type']} ({v['visual_id']})", styles["removed"]
            ))
        story.append(Spacer(1, 0.3*cm))

    # ── AI Summary ────────────────────────────────────────────
    if ai_summary and ai_summary.get("executive_summary"):
        story.append(PageBreak())
        story.append(Paragraph("AI Analysis Summary", styles["h1"]))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#2a3348"), spaceAfter=8))
        story.append(Paragraph(ai_summary["executive_summary"], styles["body"]))
        story.append(Spacer(1, 0.4*cm))

        if ai_summary.get("differences"):
            story.append(Paragraph("Differences", styles["h2"]))
            diff_data = [["Category", "What Changed", "Old", "New", "Impact"]]
            for d in ai_summary["differences"]:
                diff_data.append([
                    d.get("category", ""),
                    d.get("what_changed", "")[:80],
                    d.get("old_value", "")[:40],
                    d.get("new_value", "")[:40],
                    d.get("impact", ""),
                ])
            story.append(_make_table(diff_data, col_widths=[3*cm, 5*cm, 3*cm, 3*cm, 2*cm]))

        if ai_summary.get("risk_flags"):
            story.append(Spacer(1, 0.4*cm))
            story.append(Paragraph("Risk Flags", styles["h2"]))
            for r in ai_summary["risk_flags"]:
                story.append(Paragraph(f"⚠ {r.get('flag', '')}: {r.get('detail', '')}", styles["removed"]))

    doc.build(story)
    print(f"[exporter] ✅ PDF saved to: {output_path}")
    return output_path


# ── Styles ────────────────────────────────────────────────────
def _build_styles() -> dict:
    base = getSampleStyleSheet()
    def s(name, **kw):
        return ParagraphStyle(name, **kw)

    PRIMARY = colors.HexColor("#2CB7BE")
    ACCENT  = colors.HexColor("#C9D91A")
    GREEN   = colors.HexColor("#2F855A")
    RED     = colors.HexColor("#DC2626")
    GRAY    = colors.HexColor("#334155")
    SOFT    = colors.HexColor("#F8FAFC")
    WHITE   = colors.white
    TEXT    = colors.HexColor("#0F172A")

    return {
        "title":       s("T",  fontSize=22, textColor=PRIMARY, spaceAfter=6,  fontName="Helvetica-Bold", alignment=TA_LEFT),
        "h1":          s("H1", fontSize=14, textColor=PRIMARY, spaceBefore=12, spaceAfter=6, fontName="Helvetica-Bold"),
        "h2":          s("H2", fontSize=11, textColor=ACCENT,  spaceBefore=8,  spaceAfter=4, fontName="Helvetica-Bold"),
        "meta":        s("M",  fontSize=8,  textColor=GRAY,  spaceAfter=2,  fontName="Helvetica"),
        "body":        s("B",  fontSize=9,  textColor=TEXT, spaceAfter=4, fontName="Helvetica", leading=14),
        "detail":      s("D",  fontSize=8,  textColor=GRAY,  spaceAfter=2,  fontName="Helvetica"),
        "code":        s("C",  fontSize=8,  textColor=TEXT,
                          backColor=colors.HexColor("#F2F9F9"), fontName="Courier",
                          spaceAfter=4, leftIndent=10, borderPad=5),
        "added":       s("A",  fontSize=8.5,textColor=GREEN, spaceAfter=2,  fontName="Helvetica-Bold"),
        "removed":     s("R",  fontSize=8.5,textColor=RED,   spaceAfter=2,  fontName="Helvetica-Bold"),
        "modified":    s("Mo", fontSize=8.5,textColor=ACCENT, spaceAfter=2,  fontName="Helvetica-Bold"),
        "measure_name":s("MN", fontSize=9,  textColor=PRIMARY,  spaceAfter=2,  fontName="Helvetica-Bold"),
        "dax_label":   s("DL", fontSize=8,  textColor=GRAY,  spaceAfter=1,  fontName="Helvetica-Oblique"),
    }


def _make_table(data: list, col_widths=None) -> Table:
    GREEN = colors.HexColor("#2F855A")
    RED   = colors.HexColor("#DC2626")
    BLUE  = colors.HexColor("#2CB7BE")
    MOD   = colors.HexColor("#C9D91A")
    HBKG  = colors.HexColor("#E6F7F8")
    GRAY  = colors.HexColor("#F8FAFC")
    BORDER= colors.HexColor("#D9E7E7")

    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), HBKG),
        ("TEXTCOLOR",   (0,0), (-1,0), BLUE),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 8),
        ("FONTSIZE",    (0,1), (-1,-1), 8),
        ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, GRAY]),
        ("GRID",        (0,0), (-1,-1), 0.4, BORDER),
        ("VALIGN",      (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING",(0,0), (-1,-1), 6),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("WORDWRAP",    (0,0), (-1,-1), True),
    ])

    # Color status cells
    for i, row in enumerate(data[1:], 1):
        for j, cell in enumerate(row):
            if isinstance(cell, str):
                if cell.startswith("+ ") or cell == "+ Added":
                    style.add("TEXTCOLOR", (j,i), (j,i), GREEN)
                    style.add("BACKGROUND", (j,i), (j,i), colors.HexColor("#ECFDF5"))
                    style.add("FONTNAME",  (j,i), (j,i), "Helvetica-Bold")
                elif cell.startswith("- ") or cell == "- Removed":
                    style.add("TEXTCOLOR", (j,i), (j,i), RED)
                    style.add("BACKGROUND", (j,i), (j,i), colors.HexColor("#FEF2F2"))
                    style.add("FONTNAME",  (j,i), (j,i), "Helvetica-Bold")
                elif cell == "~ Modified":
                    style.add("TEXTCOLOR", (j,i), (j,i), MOD)
                    style.add("BACKGROUND", (j,i), (j,i), colors.HexColor("#F7FEE7"))
                    style.add("FONTNAME",  (j,i), (j,i), "Helvetica-Bold")

    t.setStyle(style)
    return t


# ═══════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════

def export_excel(
    model_results: dict,
    report_results: dict,
    old_path: str,
    new_path: str,
    output_path: Path
) -> Path:
    """Generates a multi-sheet Excel workbook of all changes."""

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Colour palette
    C = {
        "header_bg":  "1E2535",
        "header_fg":  "FFFFFF",
        "added_bg":   "D1FAE5",
        "added_fg":   "065F46",
        "removed_bg": "FEE2E2",
        "removed_fg": "991B1B",
        "modified_bg":"FEF3C7",
        "modified_fg":"92400E",
        "alt_row":    "F8FAFC",
        "border":     "E2E8F0",
        "title_bg":   "0F172A",
        "title_fg":   "4F8EF7",
    }

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws = wb.create_sheet("Summary")
    _xl_title(ws, "Power BI Diff — Summary", C)
    ws.append([])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Old File", old_path])
    ws.append(["New File", new_path])
    ws.append([])
    _xl_header(ws, ["Category", "Metric", "Count"], C)
    rows = [
        ["SemanticModel", "Tables Added",       len(model_results.get("tables_added", []))],
        ["SemanticModel", "Tables Removed",     len(model_results.get("tables_removed", []))],
        ["SemanticModel", "Columns Added",      len(model_results.get("columns_added", []))],
        ["SemanticModel", "Columns Removed",    len(model_results.get("columns_removed", []))],
        ["SemanticModel", "Columns Modified",   len(model_results.get("columns_modified", []))],
        ["SemanticModel", "Measures Changed",   len(model_results.get("measures_modified", []))],
        ["SemanticModel", "Partitions Changed", len(model_results.get("partitions_changed", []))],
        ["Report",        "Pages Added",        len(report_results.get("pages_added", []))],
        ["Report",        "Pages Removed",      len(report_results.get("pages_removed", []))],
        ["Report",        "Pages Changed",      len(report_results.get("pages_changed", []))],
    ]
    for r in rows:
        ws.append(r)
    _xl_style_rows(ws, start_row=7, C=C)
    _xl_col_widths(ws, [18, 25, 10])

    # ── Sheet 2: Tables ───────────────────────────────────────
    ws = wb.create_sheet("Tables")
    _xl_title(ws, "Table Changes", C)
    ws.append([])
    _xl_header(ws, ["Table Name", "Status", "Auto-Generated", "Columns", "Measures"], C)
    for t in model_results.get("tables_added", []):
        is_auto = "LocalDateTable" in t["name"] or "DateTableTemplate" in t["name"]
        ws.append([
            t["name"], "Added", "Yes" if is_auto else "No",
            ", ".join(t.get("columns", [])),
            ", ".join(m["name"] for m in t.get("measures", []))
        ])
        _xl_color_last_row(ws, C["added_bg"], C["added_fg"])
    for t in model_results.get("tables_removed", []):
        is_auto = "LocalDateTable" in t["name"] or "DateTableTemplate" in t["name"]
        ws.append([t["name"], "Removed", "Yes" if is_auto else "No",
                   ", ".join(t.get("columns", [])),
                   ", ".join(t.get("measures", []))])
        _xl_color_last_row(ws, C["removed_bg"], C["removed_fg"])
    _xl_col_widths(ws, [40, 12, 16, 50, 40])

    # ── Sheet 3: Columns ──────────────────────────────────────
    ws = wb.create_sheet("Columns")
    _xl_title(ws, "Column Changes", C)
    ws.append([])
    _xl_header(ws, ["Table", "Column", "Status", "DataType", "Detail"], C)
    for c in model_results.get("columns_added", []):
        ws.append([c["table"], c["column"], "Added", c.get("dataType",""), ""])
        _xl_color_last_row(ws, C["added_bg"], C["added_fg"])
    for c in model_results.get("columns_removed", []):
        ws.append([c["table"], c["column"], "Removed", "", ""])
        _xl_color_last_row(ws, C["removed_bg"], C["removed_fg"])
    for c in model_results.get("columns_modified", []):
        detail = "; ".join(f"{k}: {v['old']} → {v['new']}" for k,v in c["changes"].items())
        ws.append([c["table"], c["column"], "Modified", "", detail])
        _xl_color_last_row(ws, C["modified_bg"], C["modified_fg"])
    _xl_col_widths(ws, [30, 30, 12, 15, 50])

    # ── Sheet 4: Measures ─────────────────────────────────────
    ws = wb.create_sheet("Measures_DAX")
    _xl_title(ws, "Measure / DAX Changes", C)
    ws.append([])
    _xl_header(ws, ["Table", "Measure", "Status", "Old DAX", "New DAX", "Format String"], C)
    for m in model_results.get("measures_modified", []):
        ws.append([
            m.get("table",""), m["measure"], m["status"].capitalize(),
            m.get("old_dax") or "", m.get("new_dax") or "",
            m.get("formatString") or ""
        ])
        bg = C["added_bg"] if m["status"]=="added" else C["removed_bg"] if m["status"]=="removed" else C["modified_bg"]
        fg = C["added_fg"] if m["status"]=="added" else C["removed_fg"] if m["status"]=="removed" else C["modified_fg"]
        _xl_color_last_row(ws, bg, fg)
    _xl_col_widths(ws, [20, 35, 12, 50, 50, 18])
    # Wrap DAX columns
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=4, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Sheet 5: Partitions ───────────────────────────────────
    ws = wb.create_sheet("Partitions_Queries")
    _xl_title(ws, "Partition / Source Query Changes", C)
    ws.append([])
    _xl_header(ws, ["Table", "Partition", "Status", "Mode", "Old Query", "New Query"], C)
    for p in model_results.get("partitions_changed", []):
        ws.append([
            p["table"], p["partition"], p["status"].capitalize(),
            p.get("mode") or p.get("new_mode",""),
            p.get("old_source") or p.get("source") or "",
            p.get("new_source") or "",
        ])
        bg = C["added_bg"] if p["status"]=="added" else C["removed_bg"] if p["status"]=="removed" else C["modified_bg"]
        fg = C["added_fg"] if p["status"]=="added" else C["removed_fg"] if p["status"]=="removed" else C["modified_fg"]
        _xl_color_last_row(ws, bg, fg)
    _xl_col_widths(ws, [20, 20, 12, 12, 60, 60])

    # ── Sheet 6: Report Pages ─────────────────────────────────
    ws = wb.create_sheet("Report_Pages")
    _xl_title(ws, "Report Page Changes", C)
    ws.append([])
    _xl_header(ws, ["Page", "Status", "Visual ID", "Visual Type", "Change Type", "Fields Removed", "Fields Added", "Chart Old", "Chart New"], C)

    for pg in report_results.get("pages_changed", []):
        for v in pg.get("visuals_changed", []):
            ch = v.get("changes", {})
            ws.append([
                pg["display_name"], "Changed",
                v["visual_id"], v["visual_type"], v["status"].capitalize(),
                ", ".join(ch.get("field_swap", {}).get("removed", [])),
                ", ".join(ch.get("field_swap", {}).get("added", [])),
                ch.get("visual_type", {}).get("old", ""),
                ch.get("visual_type", {}).get("new", ""),
            ])
            _xl_color_last_row(ws, C["modified_bg"], C["modified_fg"])

    for pg in report_results.get("pages_added", []):
        for v in pg.get("visuals", []):
            ws.append([pg["display_name"], "Page Added", v["visual_id"], v["visual_type"],
                       "New Visual", "", ", ".join(v.get("fields",[])), "", ""])
            _xl_color_last_row(ws, C["added_bg"], C["added_fg"])

    for pg in report_results.get("pages_removed", []):
        for v in pg.get("visuals", []):
            ws.append([pg["display_name"], "Page Removed", v["visual_id"], v["visual_type"],
                       "Removed Visual", ", ".join(v.get("fields",[])), "", "", ""])
            _xl_color_last_row(ws, C["removed_bg"], C["removed_fg"])

    _xl_col_widths(ws, [20, 14, 30, 20, 16, 35, 35, 22, 22])

    wb.save(output_path)
    print(f"[exporter] ✅ Excel saved to: {output_path}")
    return output_path


# ── Excel Helpers ─────────────────────────────────────────────
def _xl_title(ws, text: str, C: dict):
    ws.append([text])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font  = Font(bold=True, size=13, color=C["title_fg"])
    cell.fill  = PatternFill("solid", fgColor=C["title_bg"])
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _xl_header(ws, headers: list, C: dict):
    ws.append(headers)
    row = ws.max_row
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color=C["header_fg"], size=9)
        cell.fill      = PatternFill("solid", fgColor=C["header_bg"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = _thin_border(C["border"])


def _xl_color_last_row(ws, bg: str, fg: str):
    row = ws.max_row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.font      = Font(color=fg, size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border    = _thin_border("E2E8F0")


def _xl_style_rows(ws, start_row: int, C: dict):
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        bg = "FFFFFF" if i % 2 == 0 else C["alt_row"]
        for cell in row:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="top")
            cell.border    = _thin_border(C["border"])


def _xl_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _thin_border(color: str = "E2E8F0") -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)